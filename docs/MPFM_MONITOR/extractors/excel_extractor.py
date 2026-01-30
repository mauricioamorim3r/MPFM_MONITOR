"""
MPFM Monitor - Excel Extractor
Extrator robusto para Daily Reports (Oil, Gas, Water, GasBalance)
Usa estratégia anchor-based conforme PRD de Ingestão.
"""
import os
import re
import hashlib
import sqlite3
import logging
from datetime import datetime, date
from pathlib import Path
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Generator
from enum import Enum

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl é necessário. Instale com: pip install openpyxl")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ============================================================================
# CONSTANTES E CONFIGURAÇÃO
# ============================================================================

class FileType(Enum):
    DAILY_OIL = "DAILY_OIL"
    DAILY_GAS = "DAILY_GAS"
    DAILY_WATER = "DAILY_WATER"
    GAS_BALANCE = "GAS_BALANCE"
    UNKNOWN = "UNKNOWN"


class BlockType(Enum):
    CUMULATIVE = "CUMULATIVE"
    DAY = "DAY"
    AVG = "AVG"


# Âncoras para localizar blocos (case-insensitive)
BLOCK_ANCHORS = {
    BlockType.CUMULATIVE: ["cumulative totals", "cumulative totals @ day close"],
    BlockType.DAY: ["day totals"],
    BlockType.AVG: ["flow weighted average", "flow weighted averages"]
}

# Âncoras alternativas para formatos simplificados
SIMPLE_ANCHORS = {
    BlockType.DAY: [
        "gross standard volume", "net standard volume", 
        "gross volume", "mass", "production",
        "quality parameters", "flow time"
    ]
}

# Padrão de TAG (ex: 20FT2303, 44FT0203A)
TAG_PATTERN = re.compile(r'^\d{2}[A-Z]{2}\d{4}[A-B]?$')

# Mapeamento de variáveis (nome original -> código padronizado)
VARIABLE_MAPPING = {
    # Volumes
    "gross volume": "gross_volume_m3",
    "gross standard volume": "gross_std_volume_sm3",
    "standard volume": "std_volume_sm3",
    "net standard volume": "net_std_volume_sm3",
    # Massa
    "mass": "mass_t",
    # Energia (gás)
    "energy": "energy_gj",
    "heating value": "heating_value_mj_sm3",
    # Tempo
    "flow time": "flow_time_min",
    # Pressão
    "pressure": "pressure_kpag",
    "differential pressure": "dp_kpa",
    # Temperatura
    "temperature": "temperature_c",
    # Densidade
    "line density": "line_density_kg_m3",
    "standard density": "std_density_kg_sm3",
    # Fatores
    "k-factor": "k_factor_pls_m3",
    "m factor": "m_factor",
    "meter factor": "meter_factor",
    "ctl": "ctl",
    "cpl": "cpl",
    # Qualidade
    "bs&w analyzer": "bsw_pctvol",
    "bsw": "bsw_pctvol",
}


# ============================================================================
# DATA CLASSES
# ============================================================================

@dataclass
class ReportMetadata:
    """Metadados extraídos do cabeçalho do relatório."""
    report_date: date
    period_start: Optional[datetime]
    period_end: Optional[datetime]
    report_generated_at: Optional[datetime]
    field_name: str
    source_file: str
    file_hash: str


@dataclass
class ExtractedValue:
    """Um valor extraído no formato normalizado."""
    tag: str
    section_name: str
    block_type: BlockType
    variable_raw: str
    variable_code: str
    value: Optional[float]
    unit: str
    source_sheet: str
    source_cell: str


@dataclass
class GasBalanceLine:
    """Uma linha do balanço de gás."""
    line_order: int
    line_sign: str  # '+', '-', 'TOTAL'
    description: str
    flowrate_value: Optional[float]
    flowrate_unit: str
    pd_value: Optional[float]
    pd_unit: str
    source_cell: str


@dataclass
class ExtractionResult:
    """Resultado completo da extração."""
    success: bool
    file_type: FileType
    metadata: Optional[ReportMetadata]
    values: List[ExtractedValue] = field(default_factory=list)
    gas_balance: List[GasBalanceLine] = field(default_factory=list)
    meters_found: List[str] = field(default_factory=list)
    sections_found: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


# ============================================================================
# CLASSE PRINCIPAL: ExcelExtractor
# ============================================================================

class ExcelExtractor:
    """
    Extrator de dados de relatórios diários Excel.
    Suporta: Daily_Oil, Daily_Gas, Daily_Water, GasBalance
    """
    
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        self.workbook = None
        self._file_hash = None
        self._file_type = None
    
    @property
    def file_hash(self) -> str:
        """Calcula hash SHA-256 do arquivo."""
        if self._file_hash is None:
            with open(self.file_path, 'rb') as f:
                self._file_hash = hashlib.sha256(f.read()).hexdigest()
        return self._file_hash
    
    @property
    def file_type(self) -> FileType:
        """Detecta o tipo de arquivo."""
        if self._file_type is None:
            self._file_type = self._detect_file_type()
        return self._file_type
    
    def _detect_file_type(self) -> FileType:
        """Detecta tipo de arquivo pelo nome ou conteúdo."""
        name = self.file_path.name.lower()
        
        # Detecção por nome de arquivo
        if 'daily_oil' in name:
            return FileType.DAILY_OIL
        elif 'daily_gas' in name:
            return FileType.DAILY_GAS
        elif 'daily_water' in name:
            return FileType.DAILY_WATER
        elif 'gasbalance' in name or 'gas_balance' in name:
            return FileType.GAS_BALANCE
        
        # Detecção por nome de aba
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            for sheet_name in wb.sheetnames:
                sn = sheet_name.lower()
                if sn.startswith('oil_'):
                    return FileType.DAILY_OIL
                elif sn.startswith('gas_'):
                    return FileType.DAILY_GAS
                elif sn.startswith('water_'):
                    return FileType.DAILY_WATER
                elif sn == '0001':
                    return FileType.GAS_BALANCE
            wb.close()
        except Exception:
            pass
        
        return FileType.UNKNOWN
    
    def extract(self) -> ExtractionResult:
        """
        Extrai todos os dados do arquivo Excel.
        
        Returns:
            ExtractionResult com todos os dados extraídos
        """
        result = ExtractionResult(
            success=False,
            file_type=self.file_type,
            metadata=None
        )
        
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = self._get_main_sheet()
            
            if sheet is None:
                result.errors.append("Não foi possível encontrar a aba principal")
                return result
            
            # Extrair metadados
            result.metadata = self._extract_metadata(sheet)
            if result.metadata is None:
                result.warnings.append("Metadados incompletos ou não encontrados")
            
            # Extrair dados conforme o tipo
            if self.file_type == FileType.GAS_BALANCE:
                result.gas_balance = self._extract_gas_balance(sheet, result)
            else:
                result.values = self._extract_daily_values(sheet, result)
            
            result.success = len(result.errors) == 0
            
        except Exception as e:
            result.errors.append(f"Erro ao processar arquivo: {str(e)}")
            logger.exception("Erro na extração")
        finally:
            if self.workbook:
                self.workbook.close()
        
        return result
    
    def _get_main_sheet(self) -> Optional[Worksheet]:
        """Encontra a aba principal de dados."""
        if not self.workbook:
            return None
        
        # Padrões de nome de aba por tipo
        patterns = {
            FileType.DAILY_OIL: ['oil_'],
            FileType.DAILY_GAS: ['gas_'],
            FileType.DAILY_WATER: ['water_'],
            FileType.GAS_BALANCE: ['0001', 'balance']
        }
        
        preferred = patterns.get(self.file_type, [])
        
        for sheet_name in self.workbook.sheetnames:
            name_lower = sheet_name.lower()
            for pattern in preferred:
                if pattern in name_lower or name_lower.startswith(pattern):
                    return self.workbook[sheet_name]
        
        # Fallback: primeira aba
        return self.workbook.active
    
    def _extract_metadata(self, sheet: Worksheet) -> Optional[ReportMetadata]:
        """Extrai metadados do cabeçalho do relatório."""
        metadata = {
            'report_generated_at': None,
            'field_name': None,
            'period_start': None,
            'period_end': None,
            'report_date': None
        }
        
        # Busca nas primeiras 25 linhas
        for row in range(1, 26):
            for col in range(1, 15):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    continue
                
                cell_str = str(cell.value).strip().lower()
                
                # Date and time:
                if 'date and time' in cell_str:
                    for offset in [1, 2]:
                        next_cell = sheet.cell(row=row, column=col + offset)
                        if next_cell.value:
                            metadata['report_generated_at'] = self._parse_datetime(next_cell.value)
                            break
                
                # Field:
                elif cell_str == 'field:' or cell_str.startswith('field'):
                    for offset in [1, 2]:
                        next_cell = sheet.cell(row=row, column=col + offset)
                        if next_cell.value:
                            metadata['field_name'] = str(next_cell.value).strip()
                            break
                
                # Period:
                elif 'period:' in cell_str or cell_str == 'period':
                    for offset in [1, 2]:
                        next_cell = sheet.cell(row=row, column=col + offset)
                        if next_cell.value:
                            period = self._parse_period(str(next_cell.value))
                            if period:
                                metadata['period_start'], metadata['period_end'] = period
                            break
        
        # Derivar report_date
        if metadata['period_end']:
            metadata['report_date'] = metadata['period_end'].date()
        else:
            # Tentar extrair do nome do arquivo
            date_match = re.search(r'(\d{4}-\d{2}-\d{2})', self.file_path.name)
            if date_match:
                metadata['report_date'] = datetime.strptime(date_match.group(1), "%Y-%m-%d").date()
            else:
                metadata['report_date'] = date.today()
        
        return ReportMetadata(
            report_date=metadata['report_date'],
            period_start=metadata['period_start'],
            period_end=metadata['period_end'],
            report_generated_at=metadata['report_generated_at'],
            field_name=metadata['field_name'] or "Unknown",
            source_file=self.file_path.name,
            file_hash=self.file_hash
        )
    
    def _parse_datetime(self, value: Any) -> Optional[datetime]:
        """Converte valor para datetime."""
        if isinstance(value, datetime):
            return value
        
        if value is None:
            return None
        
        str_value = str(value).strip()
        
        formats = [
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%d-%m-%Y %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M",
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(str_value, fmt)
            except ValueError:
                continue
        
        return None
    
    def _parse_period(self, period_str: str) -> Optional[Tuple[datetime, datetime]]:
        """Parse string de período como '26-Jan-2026 00:00 till 27-01-2026 00:00'."""
        try:
            parts = re.split(r'\s+till\s+', period_str, flags=re.IGNORECASE)
            if len(parts) != 2:
                return None
            
            formats = [
                "%d-%b-%Y %H:%M",
                "%d-%m-%Y %H:%M",
                "%Y-%m-%d %H:%M",
                "%d/%m/%Y %H:%M"
            ]
            
            start_dt = None
            end_dt = None
            
            for fmt in formats:
                if start_dt is None:
                    try:
                        start_dt = datetime.strptime(parts[0].strip(), fmt)
                    except ValueError:
                        pass
                if end_dt is None:
                    try:
                        end_dt = datetime.strptime(parts[1].strip(), fmt)
                    except ValueError:
                        pass
            
            if start_dt and end_dt:
                return (start_dt, end_dt)
        except Exception:
            pass
        
        return None
    
    def _extract_daily_values(self, sheet: Worksheet, result: ExtractionResult) -> List[ExtractedValue]:
        """Extrai valores dos blocos Cumulative/Day/Average."""
        values = []
        
        # Encontrar todas as âncoras
        anchors = self._find_block_anchors(sheet)
        
        logger.info(f"Âncoras encontradas: {[(bt.value, rows) for bt, rows in anchors.items() if rows]}")
        
        # Se não encontrou âncoras padrão, tentar extração simplificada
        total_anchors = sum(len(rows) for rows in anchors.values())
        if total_anchors == 0:
            logger.info("Usando extração simplificada (formato alternativo)")
            return self._extract_simple_format(sheet, result)
        
        for block_type, anchor_rows in anchors.items():
            for anchor_row in anchor_rows:
                block_values = self._extract_block(sheet, anchor_row, block_type, result)
                values.extend(block_values)
        
        return values
    
    def _extract_simple_format(self, sheet: Worksheet, result: ExtractionResult) -> List[ExtractedValue]:
        """
        Extrai valores de formato simplificado (sem âncoras padrão).
        Formato: Label | Value
        Ex: "Run 1 (20FT2303)" | "1166.91 m³"
        """
        values = []
        current_section = ""
        
        for row in range(1, sheet.max_row + 1):
            cell_a = sheet.cell(row=row, column=1)
            cell_b = sheet.cell(row=row, column=2)
            
            if cell_a.value is None:
                continue
            
            label = str(cell_a.value).strip()
            value_str = str(cell_b.value).strip() if cell_b.value else ""
            
            # Detectar seção (labels em maiúsculas ou com keywords)
            label_upper = label.upper()
            if any(kw in label_upper for kw in ['VOLUME', 'MASS', 'QUALITY', 'PARAMETERS', 'TOTALS', 'PRODUCTION']):
                current_section = label
                continue
            
            # Extrair TAG do label (formato "Run N (TAG)" ou "TAG")
            tag_match = re.search(r'\((\d{2}[A-Z]{2}\d{4}[A-B]?)\)', label)
            if not tag_match:
                tag_match = re.search(TAG_PATTERN, label)
            
            tag = tag_match.group(1) if tag_match else None
            
            # Extrair valor numérico e unidade
            value_match = re.match(r'([\d\.\,\-]+)\s*(.*)$', value_str)
            if value_match:
                num_value = self._parse_numeric(value_match.group(1))
                unit = value_match.group(2).strip()
            else:
                num_value = self._parse_numeric(value_str)
                unit = ""
            
            if num_value is not None:
                # Determinar código da variável
                variable_code = self._normalize_variable(label if not tag else current_section)
                
                # Usar label como tag se não encontrou TAG específica
                display_tag = tag if tag else label.replace(" ", "_")[:20]
                
                if tag:
                    if tag not in result.meters_found:
                        result.meters_found.append(tag)
                
                values.append(ExtractedValue(
                    tag=display_tag,
                    section_name=current_section,
                    block_type=BlockType.DAY,
                    variable_raw=label,
                    variable_code=variable_code,
                    value=num_value,
                    unit=unit,
                    source_sheet=sheet.title,
                    source_cell=f"{sheet.title}!{cell_b.coordinate}"
                ))
        
        return values
    
    def _find_block_anchors(self, sheet: Worksheet) -> Dict[BlockType, List[int]]:
        """Encontra todas as âncoras de blocos na planilha."""
        anchors = {bt: [] for bt in BlockType}
        
        for row in range(1, sheet.max_row + 1):
            for col in range(1, min(sheet.max_column + 1, 20)):
                cell = sheet.cell(row=row, column=col)
                if cell.value is None:
                    continue
                
                cell_str = str(cell.value).strip().lower()
                
                for block_type, patterns in BLOCK_ANCHORS.items():
                    for pattern in patterns:
                        if pattern in cell_str:
                            anchors[block_type].append(row)
                            break
        
        return anchors
    
    def _extract_block(self, sheet: Worksheet, anchor_row: int, block_type: BlockType, 
                       result: ExtractionResult) -> List[ExtractedValue]:
        """Extrai dados de um bloco específico."""
        values = []
        
        # Encontrar linha de TAGs (geralmente 1-5 linhas abaixo da âncora)
        tag_row = self._find_tag_row(sheet, anchor_row)
        if tag_row is None:
            result.warnings.append(f"TAG row não encontrada para âncora na linha {anchor_row}")
            return values
        
        # Mapear colunas para TAGs
        tag_columns = self._get_tag_columns(sheet, tag_row)
        if not tag_columns:
            result.warnings.append(f"Nenhuma TAG encontrada na linha {tag_row}")
            return values
        
        # Registrar TAGs encontradas
        for tag in tag_columns.values():
            if tag not in result.meters_found:
                result.meters_found.append(tag)
        
        # Encontrar section_name (nome da seção acima da âncora)
        section_names = self._get_section_names(sheet, anchor_row, tag_columns)
        for sn in section_names.values():
            if sn and sn not in result.sections_found:
                result.sections_found.append(sn)
        
        # Encontrar coluna de unidades (geralmente coluna antes das TAGs)
        unit_col = self._find_unit_column(sheet, tag_row, tag_columns)
        
        # Ler linhas de variáveis
        current_row = tag_row + 1
        max_empty_rows = 3
        empty_count = 0
        
        while current_row <= sheet.max_row and empty_count < max_empty_rows:
            # Verificar se chegou em outra âncora
            first_cell = sheet.cell(row=current_row, column=1).value
            if first_cell:
                first_str = str(first_cell).strip().lower()
                if any(p in first_str for patterns in BLOCK_ANCHORS.values() for p in patterns):
                    break
            
            # Encontrar nome da variável (procurar nas primeiras colunas)
            variable_raw = None
            var_col = None
            for col in range(1, min(list(tag_columns.keys())[0] if tag_columns else 10, 10)):
                cell = sheet.cell(row=current_row, column=col)
                if cell.value and str(cell.value).strip():
                    val = str(cell.value).strip()
                    # Ignorar células que parecem números ou unidades
                    if not re.match(r'^[\d\.\,\-]+$', val) and len(val) > 2:
                        variable_raw = val
                        var_col = col
                        break
            
            if variable_raw is None:
                empty_count += 1
                current_row += 1
                continue
            
            empty_count = 0
            
            # Obter unidade
            unit = ""
            if unit_col:
                unit_cell = sheet.cell(row=current_row, column=unit_col)
                if unit_cell.value:
                    unit = str(unit_cell.value).strip()
            
            # Padronizar nome da variável
            variable_code = self._normalize_variable(variable_raw)
            
            # Extrair valores para cada TAG
            for col, tag in tag_columns.items():
                cell = sheet.cell(row=current_row, column=col)
                value = self._parse_numeric(cell.value)
                
                section_name = section_names.get(col, "")
                
                values.append(ExtractedValue(
                    tag=tag,
                    section_name=section_name,
                    block_type=block_type,
                    variable_raw=variable_raw,
                    variable_code=variable_code,
                    value=value,
                    unit=unit,
                    source_sheet=sheet.title,
                    source_cell=f"{sheet.title}!{cell.coordinate}"
                ))
            
            current_row += 1
        
        return values
    
    def _find_tag_row(self, sheet: Worksheet, anchor_row: int) -> Optional[int]:
        """Encontra a linha contendo TAGs abaixo de uma âncora."""
        for offset in range(1, 8):
            row = anchor_row + offset
            if row > sheet.max_row:
                break
            
            tag_count = 0
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value and TAG_PATTERN.match(str(cell.value).strip()):
                    tag_count += 1
            
            if tag_count >= 2:
                return row
        
        return None
    
    def _get_tag_columns(self, sheet: Worksheet, tag_row: int) -> Dict[int, str]:
        """Mapeia coluna -> TAG."""
        tags = {}
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=tag_row, column=col)
            if cell.value:
                tag = str(cell.value).strip()
                if TAG_PATTERN.match(tag):
                    tags[col] = tag
        return tags
    
    def _get_section_names(self, sheet: Worksheet, anchor_row: int, 
                           tag_columns: Dict[int, str]) -> Dict[int, str]:
        """Encontra nomes de seção para cada coluna de TAG."""
        sections = {}
        
        # Procurar 1-5 linhas acima da âncora
        for col in tag_columns.keys():
            section_name = ""
            for offset in range(1, 6):
                row = anchor_row - offset
                if row < 1:
                    break
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    val = str(cell.value).strip()
                    # Ignorar se parece ser TAG ou número
                    if not TAG_PATTERN.match(val) and not re.match(r'^[\d\.\,]+$', val):
                        section_name = val
                        break
            sections[col] = section_name
        
        return sections
    
    def _find_unit_column(self, sheet: Worksheet, tag_row: int, 
                          tag_columns: Dict[int, str]) -> Optional[int]:
        """Encontra a coluna de unidades."""
        if not tag_columns:
            return None
        
        first_tag_col = min(tag_columns.keys())
        
        # Procurar coluna com unidades típicas antes das TAGs
        for col in range(first_tag_col - 1, 0, -1):
            cell = sheet.cell(row=tag_row + 1, column=col)
            if cell.value:
                val = str(cell.value).strip().lower()
                if any(u in val for u in ['m³', 'sm³', 'kpa', '°c', 'kg', 't', 'min', 'gj', '%']):
                    return col
        
        return first_tag_col - 1 if first_tag_col > 1 else None
    
    def _normalize_variable(self, raw: str) -> str:
        """Normaliza nome de variável para código padronizado."""
        raw_clean = raw.strip().lower()
        
        # Remover caracteres especiais
        raw_clean = re.sub(r'[^\w\s]', '', raw_clean)
        
        # Buscar no mapeamento
        for pattern, code in VARIABLE_MAPPING.items():
            if pattern in raw_clean:
                return code
        
        # Fallback: converter para snake_case
        return re.sub(r'\s+', '_', raw_clean)
    
    def _parse_numeric(self, value: Any) -> Optional[float]:
        """Converte valor para float."""
        if value is None:
            return None
        
        if isinstance(value, (int, float)):
            return float(value)
        
        str_value = str(value).strip()
        
        # Valores nulos
        if str_value in ('-', '', 'None', 'null', 'N/A', 'n/a', '#N/A', '#REF!'):
            return None
        
        try:
            # Substituir vírgula por ponto
            str_value = str_value.replace(',', '.')
            return float(str_value)
        except ValueError:
            return None
    
    def _extract_gas_balance(self, sheet: Worksheet, result: ExtractionResult) -> List[GasBalanceLine]:
        """Extrai dados do balanço de gás."""
        lines = []
        
        # Encontrar início da tabela de balanço
        balance_row = None
        for row in range(1, sheet.max_row + 1):
            for col in range(1, 10):
                cell = sheet.cell(row=row, column=col)
                if cell.value and 'gas balance' in str(cell.value).lower():
                    balance_row = row
                    break
            if balance_row:
                break
        
        if balance_row is None:
            result.errors.append("Tabela 'Gas balance' não encontrada")
            return lines
        
        # Encontrar colunas (Sign, Description, Flow rate unit, Flow rate value, PD unit, PD value)
        header_row = balance_row + 1
        columns = {}
        
        for col in range(1, 20):
            cell = sheet.cell(row=header_row, column=col)
            if cell.value:
                val = str(cell.value).strip().lower()
                if 'sign' in val or val in ['+', '-']:
                    columns['sign'] = col
                elif 'flow rate' in val and 'unit' not in val:
                    columns['flowrate_value'] = col
                elif 'pd' in val and 'unit' not in val:
                    columns['pd_value'] = col
        
        # Se não encontrou pelas headers, usar posição padrão
        if not columns:
            # Estrutura típica: Sign, Description, Flow unit, Flow value, PD unit, PD value
            columns = {
                'sign': 1,
                'description': 2,
                'flowrate_unit': 3,
                'flowrate_value': 4,
                'pd_unit': 5,
                'pd_value': 6
            }
        
        # Ler linhas
        current_row = header_row + 1
        line_order = 0
        
        while current_row <= sheet.max_row:
            sign_cell = sheet.cell(row=current_row, column=columns.get('sign', 1))
            
            if sign_cell.value is None:
                # Verificar se é linha de total
                for col in range(1, 10):
                    cell = sheet.cell(row=current_row, column=col)
                    if cell.value and 'total' in str(cell.value).lower():
                        sign_cell = cell
                        break
            
            if sign_cell.value is None:
                current_row += 1
                continue
            
            sign_str = str(sign_cell.value).strip()
            
            # Determinar tipo de linha
            if sign_str == '+':
                line_sign = '+'
            elif sign_str == '-':
                line_sign = '-'
            elif 'total' in sign_str.lower():
                line_sign = 'TOTAL'
            else:
                current_row += 1
                continue
            
            line_order += 1
            
            # Extrair descrição
            desc_col = columns.get('description', 2)
            desc_cell = sheet.cell(row=current_row, column=desc_col)
            description = str(desc_cell.value).strip() if desc_cell.value else ""
            
            # Extrair valores
            flowrate_value = None
            pd_value = None
            flowrate_unit = ""
            pd_unit = ""
            
            if 'flowrate_value' in columns:
                fv_cell = sheet.cell(row=current_row, column=columns['flowrate_value'])
                flowrate_value = self._parse_numeric(fv_cell.value)
            
            if 'pd_value' in columns:
                pv_cell = sheet.cell(row=current_row, column=columns['pd_value'])
                pd_value = self._parse_numeric(pv_cell.value)
            
            if 'flowrate_unit' in columns:
                fu_cell = sheet.cell(row=current_row, column=columns['flowrate_unit'])
                flowrate_unit = str(fu_cell.value).strip() if fu_cell.value else ""
            
            if 'pd_unit' in columns:
                pu_cell = sheet.cell(row=current_row, column=columns['pd_unit'])
                pd_unit = str(pu_cell.value).strip() if pu_cell.value else ""
            
            lines.append(GasBalanceLine(
                line_order=line_order,
                line_sign=line_sign,
                description=description,
                flowrate_value=flowrate_value,
                flowrate_unit=flowrate_unit,
                pd_value=pd_value,
                pd_unit=pd_unit,
                source_cell=f"{sheet.title}!{sign_cell.coordinate}"
            ))
            
            # Parar depois do TOTAL
            if line_sign == 'TOTAL':
                break
            
            current_row += 1
        
        return lines


# ============================================================================
# CLASSE: DatabaseLoader
# ============================================================================

class DatabaseLoader:
    """Carrega dados extraídos no banco de dados SQLite."""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self._ensure_schema()
    
    def _ensure_schema(self):
        """Schema já deve existir - não fazer nada."""
        pass
    
    def load(self, result: ExtractionResult, installation_id: int = 1) -> Dict:
        """
        Carrega dados extraídos no banco.
        
        Returns:
            Dict com estatísticas de carregamento
        """
        stats = {
            'import_id': None,
            'snapshot_id': None,
            'meters_created': 0,
            'sections_created': 0,
            'values_inserted': 0,
            'balance_lines_inserted': 0
        }
        
        if not result.success or not result.metadata:
            return stats
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # 1. Registrar importação
            cursor.execute("""
                INSERT OR IGNORE INTO import_log 
                (file_name, file_hash, file_type, report_date, period_start, period_end, 
                 field_name, records_extracted, status)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'SUCCESS')
            """, (
                result.metadata.source_file,
                result.metadata.file_hash,
                result.file_type.value,
                result.metadata.report_date.isoformat(),
                result.metadata.period_start.isoformat() if result.metadata.period_start else None,
                result.metadata.period_end.isoformat() if result.metadata.period_end else None,
                result.metadata.field_name,
                len(result.values) + len(result.gas_balance)
            ))
            
            cursor.execute("SELECT id FROM import_log WHERE file_hash = ?", 
                          (result.metadata.file_hash,))
            row = cursor.fetchone()
            if row:
                stats['import_id'] = row[0]
            
            # 2. Criar/obter snapshot
            cursor.execute("""
                INSERT OR IGNORE INTO daily_snapshot 
                (installation_id, report_date, period_start, period_end, report_generated_at)
                VALUES (?, ?, ?, ?, ?)
            """, (
                installation_id,
                result.metadata.report_date.isoformat(),
                result.metadata.period_start.isoformat() if result.metadata.period_start else None,
                result.metadata.period_end.isoformat() if result.metadata.period_end else None,
                result.metadata.report_generated_at.isoformat() if result.metadata.report_generated_at else None
            ))
            
            cursor.execute("""
                SELECT id FROM daily_snapshot 
                WHERE installation_id = ? AND report_date = ?
            """, (installation_id, result.metadata.report_date.isoformat()))
            row = cursor.fetchone()
            if row:
                stats['snapshot_id'] = row[0]
            
            # 3. Criar medidores encontrados
            fluid_type = {
                FileType.DAILY_OIL: 'OIL',
                FileType.DAILY_GAS: 'GAS',
                FileType.DAILY_WATER: 'WATER'
            }.get(result.file_type, None)
            
            for tag in result.meters_found:
                cursor.execute("""
                    INSERT OR IGNORE INTO meter (installation_id, tag, fluid_type)
                    VALUES (?, ?, ?)
                """, (installation_id, tag, fluid_type))
                if cursor.rowcount > 0:
                    stats['meters_created'] += 1
            
            # 4. Criar seções encontradas
            for section_name in result.sections_found:
                if section_name:
                    cursor.execute("""
                        INSERT OR IGNORE INTO section (installation_id, name, fluid_type)
                        VALUES (?, ?, ?)
                    """, (installation_id, section_name, fluid_type))
                    if cursor.rowcount > 0:
                        stats['sections_created'] += 1
            
            # 5. Inserir valores
            for val in result.values:
                # Obter IDs
                cursor.execute("SELECT id FROM meter WHERE tag = ? AND installation_id = ?", 
                              (val.tag, installation_id))
                meter_row = cursor.fetchone()
                meter_id = meter_row[0] if meter_row else None
                
                cursor.execute("SELECT id FROM section WHERE name = ? AND installation_id = ?", 
                              (val.section_name, installation_id))
                section_row = cursor.fetchone()
                section_id = section_row[0] if section_row else None
                
                cursor.execute("""
                    INSERT OR REPLACE INTO daily_measurement
                    (snapshot_id, import_id, meter_id, section_id, block_type, 
                     variable_code, variable_raw, value, unit, source_sheet, source_cell)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    stats['snapshot_id'],
                    stats['import_id'],
                    meter_id,
                    section_id,
                    val.block_type.value,
                    val.variable_code,
                    val.variable_raw,
                    val.value,
                    val.unit,
                    val.source_sheet,
                    val.source_cell
                ))
                stats['values_inserted'] += 1
            
            # 6. Inserir balanço de gás
            for line in result.gas_balance:
                cursor.execute("""
                    INSERT INTO gas_balance
                    (snapshot_id, import_id, line_order, line_sign, line_description,
                     flowrate_value, flowrate_unit, pd_value, pd_unit, source_cell)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    stats['snapshot_id'],
                    stats['import_id'],
                    line.line_order,
                    line.line_sign,
                    line.description,
                    line.flowrate_value,
                    line.flowrate_unit,
                    line.pd_value,
                    line.pd_unit,
                    line.source_cell
                ))
                stats['balance_lines_inserted'] += 1
            
            conn.commit()
            
        except Exception as e:
            conn.rollback()
            logger.exception("Erro ao carregar dados no banco")
            raise
        finally:
            conn.close()
        
        return stats


# ============================================================================
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO
# ============================================================================

def process_excel_file(file_path: str, db_path: str, installation_id: int = 1) -> Dict:
    """
    Processa um arquivo Excel e carrega no banco de dados.
    
    Args:
        file_path: Caminho do arquivo Excel
        db_path: Caminho do banco SQLite
        installation_id: ID da instalação
        
    Returns:
        Dict com resultado do processamento
    """
    logger.info(f"Processando: {file_path}")
    
    extractor = ExcelExtractor(file_path)
    result = extractor.extract()
    
    logger.info(f"Tipo: {result.file_type.value}")
    logger.info(f"Data: {result.metadata.report_date if result.metadata else 'N/A'}")
    logger.info(f"Valores extraídos: {len(result.values)}")
    logger.info(f"Linhas balanço: {len(result.gas_balance)}")
    logger.info(f"TAGs encontradas: {result.meters_found}")
    
    if result.warnings:
        for w in result.warnings:
            logger.warning(w)
    
    if result.errors:
        for e in result.errors:
            logger.error(e)
        return {'success': False, 'errors': result.errors}
    
    # Carregar no banco
    loader = DatabaseLoader(db_path)
    stats = loader.load(result, installation_id)
    
    logger.info(f"Carregamento concluído: {stats}")
    
    return {
        'success': True,
        'file_type': result.file_type.value,
        'report_date': result.metadata.report_date.isoformat() if result.metadata else None,
        'extraction': {
            'values': len(result.values),
            'balance_lines': len(result.gas_balance),
            'meters_found': result.meters_found,
            'sections_found': result.sections_found
        },
        'loading': stats
    }


def process_directory(dir_path: str, db_path: str, installation_id: int = 1) -> List[Dict]:
    """
    Processa todos os arquivos Excel em um diretório.
    
    Args:
        dir_path: Diretório com arquivos Excel
        db_path: Caminho do banco SQLite
        installation_id: ID da instalação
        
    Returns:
        Lista de resultados por arquivo
    """
    results = []
    dir_path = Path(dir_path)
    
    for file_path in dir_path.glob("*.xlsx"):
        if file_path.name.startswith("~$"):  # Ignorar arquivos temporários
            continue
        
        try:
            result = process_excel_file(str(file_path), db_path, installation_id)
            results.append({'file': file_path.name, **result})
        except Exception as e:
            logger.exception(f"Erro ao processar {file_path.name}")
            results.append({'file': file_path.name, 'success': False, 'error': str(e)})
    
    return results


# ============================================================================
# CLI
# ============================================================================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: python excel_extractor.py <arquivo.xlsx ou diretório> [banco.db]")
        sys.exit(1)
    
    path = sys.argv[1]
    db_path = sys.argv[2] if len(sys.argv) > 2 else "mpfm_monitor.db"
    
    if os.path.isdir(path):
        results = process_directory(path, db_path)
        print(f"\nProcessados {len(results)} arquivos")
        for r in results:
            status = "✅" if r.get('success') else "❌"
            print(f"  {status} {r['file']}")
    else:
        result = process_excel_file(path, db_path)
        print(f"\nResultado: {'✅ Sucesso' if result.get('success') else '❌ Falha'}")
